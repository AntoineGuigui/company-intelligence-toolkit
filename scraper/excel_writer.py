"""
Excel writer — appends extracted company data to DataBase.xlsm.
"""
import logging
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from .config import EXCEL_COLUMNS

logger = logging.getLogger(__name__)

def _map_extraction_to_row(data: dict) -> dict:
    return {
        "Company Name": data.get("company_name", ""),
        "Country": data.get("country", ""),
        "Field": data.get("field", ""),
        "Activity": data.get("activity", ""),
        "Locations": data.get("locations", ""),
        "Founded": data.get("founded", ""),
        "N° employees": data.get("employees", ""),
        "Key people": data.get("key_people", ""),
        "Type Ownership": data.get("type_ownership", ""),
        "Confidence Index": "1",
        "Business Overview": data.get("business_overview", ""),
        "Business relationships": data.get("business_relationships", ""),
        "Capability": data.get("capability", ""),
        "Notes": data.get("notes", ""),
    }

def write_to_excel(data: dict, excel_path: str = "DataBase.xlsm") -> bool:
    row = _map_extraction_to_row(data)
    path = Path(excel_path)
    try:
        if path.exists():
            wb = load_workbook(path, keep_vba=True)
            ws = wb.active
            next_row = ws.max_row + 1
            for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
                ws.cell(row=next_row, column=col_idx, value=row.get(col_name, ""))
            wb.save(path)
        else:
            df = pd.DataFrame([row], columns=EXCEL_COLUMNS)
            df.to_excel(path, index=False, engine="openpyxl")
        return True
    except Exception as e:
        logger.error(f"Failed to write {row.get('Company Name', '?')}: {e}")
        return False

def write_batch(companies: list[dict], excel_path: str = "DataBase.xlsm") -> int:
    return sum(1 for d in companies if write_to_excel(d, excel_path))
